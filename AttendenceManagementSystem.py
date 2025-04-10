import tkinter as tk

import os
import datetime
import openpyxl
from tkinter import ttk, simpledialog, messagebox, filedialog
from datetime import datetime
from openpyxl import Workbook, load_workbook
import os


class AttendanceApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Attendance Management System")
        self.geometry("1200x800")
        self.configure(bg="#f0f0f0")
        
        # Initialize data structures
        self.name_data = []  # Stores (name, email, sap) tuples
        self.attendance_data = {}  # Maps names to attendance vars
        self.settings = {
            "default_month": datetime.now().strftime("%B"),
            "default_save_path": os.path.expanduser("~/Documents"),
            "theme": "light"
        }
        
        # Create container frame
        self.container = tk.Frame(self)
        self.container.pack(side="top", fill="both", expand=True)
        self.container.grid_rowconfigure(0, weight=1)
        self.container.grid_columnconfigure(0, weight=1)
        
        # Initialize all frames
        self.frames = {}
        for F in (MainMenu, AttendancePage, ReportsPage, SettingsPage):
            frame = F(parent=self.container, controller=self)
            self.frames[F.__name__] = frame
            frame.grid(row=0, column=0, sticky="nsew")
        
        # Configure style
        self.configure_styles()
        
        # Show the main menu first
        self.show_frame("MainMenu")
        
    def show_frame(self, page_name):
        """Show a frame for the given page name"""
        frame = self.frames[page_name]
        frame.tkraise()
        if hasattr(frame, "on_show"):
            frame.on_show()
        
    def configure_styles(self):
        """Configure the application styles"""
        self.style = ttk.Style()
        self.style.configure('TFrame', background="#f0f0f0")
        self.style.configure('TLabel', background="#f0f0f0", font=('Arial', 10))
        self.style.configure('Header.TLabel', font=('Arial', 12, 'bold'))
        self.style.configure('Title.TLabel', font=('Arial', 16, 'bold'))
        self.style.configure('TButton', font=('Arial', 10), padding=5)
        self.style.configure('Menu.TButton', font=('Arial', 12), padding=10)
        self.style.configure('Summary.TLabel', font=('Arial', 10, 'bold'), foreground='blue')
        
    def get_attendance_summary(self):
        """Generate attendance summary data"""
        if not self.name_data:
            return []
            
        total_days = len(self.attendance_data[self.name_data[0][0]]) if self.name_data else 0
        summary = []
        
        for name, _, _ in self.name_data:
            present_days = sum(var.get() for var in self.attendance_data[name])
            percentage = (present_days / total_days) * 100 if total_days > 0 else 0
            summary.append({
                "name": name,
                "present": present_days,
                "total": total_days,
                "percentage": percentage
            })
        
        return summary

class MainMenu(tk.Frame):
    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)
        self.controller = controller
        self.configure(bg="#f0f0f0")
        
        # Title
        title = ttk.Label(self, text="Attendance Management System", style="Title.TLabel")
        title.pack(pady=50)
        
        # Menu buttons
        button_frame = tk.Frame(self, bg="#f0f0f0")
        button_frame.pack(expand=True)
        
        buttons = [
            ("Take Attendance", "AttendancePage"),
            ("View Reports", "ReportsPage"),
            ("Settings", "SettingsPage"),
            ("Exit", lambda: controller.destroy())
        ]
        
        for text, command in buttons:
            btn = ttk.Button(button_frame, text=text, style="Menu.TButton",
                            command=lambda c=command: self.controller.show_frame(c) if isinstance(c, str) else c())
            btn.pack(fill="x", padx=100, pady=10)

class AttendancePage(tk.Frame):
    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)
        self.controller = controller
        self.configure(bg="#f0f0f0")
        
        # Header
        header = ttk.Frame(self)
        header.pack(fill="x", pady=10)
        
        ttk.Button(header, text="← Main Menu", 
                  command=lambda: controller.show_frame("MainMenu")).pack(side="left", padx=10)
        
        self.title = ttk.Label(header, text="Attendance Recording", style="Title.TLabel")
        self.title.pack(side="left", expand=True)
        
        # Control panel
        control_frame = ttk.Frame(self, padding="10")
        control_frame.pack(fill="x")
        
        # Date and Month
        ttk.Label(control_frame, text="Date:").grid(row=0, column=0, sticky="e", padx=5)
        self.date_var = tk.StringVar(value=datetime.now().strftime("%Y-%m-%d"))
        ttk.Entry(control_frame, textvariable=self.date_var, width=12).grid(row=0, column=1, sticky="w", padx=5)
        
        ttk.Label(control_frame, text="Month:").grid(row=0, column=2, sticky="e", padx=5)
        self.month_var = tk.StringVar(value=controller.settings["default_month"])
        months = ["January", "February", "March", "April", "May", "June",
                 "July", "August", "September", "October", "November", "December"]
        ttk.Combobox(control_frame, textvariable=self.month_var, 
                    values=months, state="readonly", width=10).grid(row=0, column=3, sticky="w", padx=5)
        
        # Buttons
        button_frame = ttk.Frame(control_frame)
        button_frame.grid(row=0, column=4, columnspan=4, sticky="e", padx=10)
        
        ttk.Button(button_frame, text="Add Attendee", command=self.add_attendee).pack(side="left", padx=5)
        ttk.Button(button_frame, text="Remove Last", command=self.remove_attendee).pack(side="left", padx=5)
        ttk.Button(button_frame, text="Clear All", command=self.clear_attendees).pack(side="left", padx=5)
        ttk.Button(button_frame, text="Save", command=self.save_attendance).pack(side="left", padx=5)
        ttk.Button(button_frame, text="Load", command=self.load_attendance).pack(side="left", padx=5)
        
        # Attendees display
        self.create_attendees_display()
        
        # Summary display
        self.create_summary_display()
    
    def create_attendees_display(self):
        """Create the scrollable attendees display"""
        container = ttk.Frame(self)
        container.pack(fill="both", expand=True, padx=10, pady=10)
        
        # Canvas and scrollbars
        self.canvas = tk.Canvas(container, borderwidth=0, background="#ffffff")
        vsb = ttk.Scrollbar(container, orient="vertical", command=self.canvas.yview)
        hsb = ttk.Scrollbar(container, orient="horizontal", command=self.canvas.xview)
        self.canvas.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        
        vsb.pack(side="right", fill="y")
        hsb.pack(side="bottom", fill="x")
        self.canvas.pack(side="left", fill="both", expand=True)
        
        # Frame inside canvas
        self.attendees_frame = ttk.Frame(self.canvas)
        self.canvas.create_window((0, 0), window=self.attendees_frame, anchor="nw")
        
        # Configure resizing
        self.attendees_frame.bind("<Configure>", lambda e: self.canvas.configure(scrollregion=self.canvas.bbox("all")))
    
    def create_summary_display(self):
        """Create the summary display at the bottom"""
        summary_frame = ttk.Frame(self, padding="10")
        summary_frame.pack(fill="x")
        
        ttk.Label(summary_frame, text="Attendance Summary", style="Header.TLabel").pack(anchor="w")
        
        self.summary_text = tk.Text(summary_frame, height=6, wrap="word", 
                                  font=('Arial', 10), bg="#f9f9f9",fg="black")
        scrollbar = ttk.Scrollbar(summary_frame, command=self.summary_text.yview)
        scrollbar.pack(side="right", fill="y")
        self.summary_text.config(yscrollcommand=scrollbar.set)
        self.summary_text.pack(fill="x")
    
    def add_attendee(self):
        """Add a new attendee"""
        name = simpledialog.askstring("Add Attendee", "Enter the name:")
        if not name:
            return
            
        email = simpledialog.askstring("Add Attendee", f"Enter email for {name}:")
        sap = simpledialog.askstring("Add Attendee", f"Enter SAP ID for {name}:")
        
        self.controller.name_data.append((name, email, sap))
        attendance_vars = [tk.IntVar(value=0) for _ in range(31)]
        self.controller.attendance_data[name] = attendance_vars
        
        self.update_display()
    
    def remove_attendee(self):
        """Remove the last attendee"""
        if not self.controller.name_data:
            messagebox.showwarning("Warning", "No attendees to remove")
            return
            
        self.controller.name_data.pop()
        self.controller.attendance_data.popitem()
        self.update_display()
    
    def clear_attendees(self):
        """Clear all attendees"""
        if not self.controller.name_data:
            return
            
        if messagebox.askyesno("Confirm", "Clear all attendees?"):
            self.controller.name_data.clear()
            self.controller.attendance_data.clear()
            self.update_display()
    
    def update_display(self):
        """Update the attendees and summary displays"""
        # Clear existing widgets
        for widget in self.attendees_frame.winfo_children():
            widget.destroy()
            
        if not self.controller.name_data:
            return
            
        # Create header row
        ttk.Label(self.attendees_frame, text="Name", style="Header.TLabel"
                 ).grid(row=0, column=0, padx=5, pady=2, sticky="w")
        ttk.Label(self.attendees_frame, text="Email", style="Header.TLabel"
                 ).grid(row=0, column=1, padx=5, pady=2, sticky="w")
        ttk.Label(self.attendees_frame, text="SAP ID", style="Header.TLabel"
                 ).grid(row=0, column=2, padx=5, pady=2, sticky="w")
        
        # Day headers (1-31)
        for day in range(1, 32):
            ttk.Label(self.attendees_frame, text=str(day), style="Header.TLabel"
                     ).grid(row=0, column=2+day, padx=2, pady=2)
        
        # Add attendee rows
        for row, (name, email, sap) in enumerate(self.controller.name_data, start=1):
            ttk.Label(self.attendees_frame, text=name
                     ).grid(row=row, column=0, padx=5, pady=2, sticky="w")
            ttk.Label(self.attendees_frame, text=email
                     ).grid(row=row, column=1, padx=5, pady=2, sticky="w")
            ttk.Label(self.attendees_frame, text=sap
                     ).grid(row=row, column=2, padx=5, pady=2, sticky="w")
            
            for day, var in enumerate(self.controller.attendance_data[name], start=1):
                cb = tk.Checkbutton(self.attendees_frame, variable=var, 
                                   command=self.update_summary)
                cb.grid(row=row, column=2+day, padx=2, pady=2)
        
        self.update_summary()
    
    def update_summary(self):
        """Update the summary display"""
        self.summary_text.config(state="normal")
        self.summary_text.delete(1.0, tk.END)
        
        if not self.controller.name_data:
            self.summary_text.insert(tk.END, "No attendees added yet.")
            self.summary_text.config(state="disabled")
            return
            
        summary = self.controller.get_attendance_summary()
        for item in summary:
            self.summary_text.insert(tk.END, 
                                   f"{item['name']}: {item['present']}/{item['total']} days ({item['percentage']:.1f}%)\n")
            
        self.summary_text.config(state="disabled")
    
    def save_attendance(self):
        """Save attendance to Excel file"""
        if not self.controller.name_data:
            messagebox.showwarning("Warning", "No attendance data to save")
            return
            
        wb = Workbook()
        ws = wb.active
        
        # Add headers
        ws['A1'] = "Month"
        ws['B1'] = self.month_var.get()
        ws['A2'] = "Date of update"
        ws['B2'] = self.date_var.get()
        ws['A3'] = "Name"
        ws['B3'] = "Email"
        ws['C3'] = "SAP ID"
        
        # Day headers
        for day in range(1, 32):
            ws.cell(row=3, column=3+day, value=day)
        
        # Add attendee data
        for row, (name, email, sap) in enumerate(self.controller.name_data, start=4):
            ws.cell(row=row, column=1, value=name)
            ws.cell(row=row, column=2, value=email)
            ws.cell(row=row, column=3, value=sap)
            
            for day, var in enumerate(self.controller.attendance_data[name], start=1):
                status = "Present" if var.get() else "Absent"
                ws.cell(row=row, column=3+day, value=status)
        
        # Save file
        default_path = self.controller.settings["default_save_path"]
        file_path = filedialog.asksaveasfilename(
            initialdir=default_path,
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            title="Save Attendance Sheet"
        )
        
        if file_path:
            try:
                wb.save(file_path)
                messagebox.showinfo("Success", f"Attendance saved to:\n{file_path}")
            except Exception as e:
                messagebox.showerror("Error", f"Failed to save file:\n{str(e)}")
    
    def load_attendance(self):
        """Load attendance from Excel file"""
        default_path = self.controller.settings["default_save_path"]
        file_path = filedialog.askopenfilename(
            initialdir=default_path,
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            title="Load Attendance Sheet"
        )
        
        if not file_path:
            return
            
        try:
            wb = load_workbook(file_path)
            ws = wb.active
            
            # Clear current data
            self.controller.name_data.clear()
            self.controller.attendance_data.clear()
            
            # Load metadata
            self.month_var.set(ws['B1'].value)
            self.date_var.set(ws['B2'].value)
            
            # Load attendee data
            for row in ws.iter_rows(min_row=4, values_only=True):
                if not row[0]:
                    continue
                    
                name, email, sap = row[0], row[1], row[2]
                self.controller.name_data.append((name, email, sap))
                
                attendance_vars = []
                for status in row[3:34]:  # First 31 days
                    var = tk.IntVar(value=1 if status == "Present" else 0)
                    attendance_vars.append(var)
                self.controller.attendance_data[name] = attendance_vars
            
            self.update_display()
            messagebox.showinfo("Success", f"Loaded attendance from:\n{file_path}")
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load file:\n{str(e)}")
    
    def on_show(self):
        """Called when the frame is shown"""
        self.update_display()

class ReportsPage(tk.Frame):
    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)
        self.controller = controller
        self.configure(bg="#f0f0f0")
        
        # Header
        header = ttk.Frame(self)
        header.pack(fill="x", pady=10)
        
        ttk.Button(header, text="← Main Menu", 
                  command=lambda: controller.show_frame("MainMenu")).pack(side="left", padx=10)
        
        ttk.Label(header, text="Attendance Reports", style="Title.TLabel").pack(side="left", expand=True)
        
        # Content
        self.content = ttk.Frame(self)
        self.content.pack(fill="both", expand=True, padx=20, pady=20)
        
        # Initialize report display
        self.init_report_display()
    
    def init_report_display(self):
        """Initialize the report display area"""
        # Clear existing widgets
        for widget in self.content.winfo_children():
            widget.destroy()
        
        # Summary statistics
        summary_frame = ttk.LabelFrame(self.content, text="Summary Statistics", padding=10)
        summary_frame.pack(fill="x", pady=10)
        
        self.summary_text = tk.Text(summary_frame, height=10, wrap="word", 
                                  font=('Arial', 10), bg="#f9f9f9", fg="black")
        scrollbar = ttk.Scrollbar(summary_frame, command=self.summary_text.yview)
        scrollbar.pack(side="right", fill="y")
        self.summary_text.config(yscrollcommand=scrollbar.set)
        self.summary_text.pack(fill="both", expand=True)
        
        # Export buttons
        button_frame = ttk.Frame(self.content)
        button_frame.pack(fill="x", pady=10)
        
        ttk.Button(button_frame, text="Export to Excel", command=self.export_to_excel
                  ).pack(side="left", padx=5)
        ttk.Button(button_frame, text="Generate Report", command=self.generate_report
                  ).pack(side="left", padx=5)
        ttk.Button(button_frame, text="Print", command=self.print_report
                  ).pack(side="left", padx=5)
    
    def update_report(self):
        """Update the report display with current data"""
        self.summary_text.config(state="normal")
        self.summary_text.delete(1.0, tk.END)
        
        summary = self.controller.get_attendance_summary()
        if not summary:
            self.summary_text.insert(tk.END, "No attendance data available.")
            self.summary_text.config(state="disabled")
            return
        
        # Calculate overall statistics
        total_days = summary[0]['total'] if summary else 0
        total_present = sum(item['present'] for item in summary)
        total_possible = total_days * len(summary) if summary else 1
        overall_percentage = (total_present / total_possible) * 100 if total_possible > 0 else 0
        
        # Add header
        self.summary_text.insert(tk.END, f"ATTENDANCE REPORT\n", "header")
        self.summary_text.insert(tk.END, f"Date: {datetime.now().strftime('%Y-%m-%d')}\n")
        self.summary_text.insert(tk.END, f"Total attendees: {len(summary)}\n")
        self.summary_text.insert(tk.END, f"Total days recorded: {total_days}\n")
        self.summary_text.insert(tk.END, f"Overall attendance: {overall_percentage:.1f}%\n\n")
        
        # Add individual records
        self.summary_text.insert(tk.END, "INDIVIDUAL RECORDS:\n", "header")
        for item in summary:
            self.summary_text.insert(tk.END, 
                                   f"{item['name']}: {item['present']}/{item['total']} days ({item['percentage']:.1f}%)\n")
        
        self.summary_text.tag_configure("header", font=('Arial', 10, 'bold'))
        self.summary_text.config(state="disabled")
    
    def export_to_excel(self):
        """Export the report to Excel"""
        summary = self.controller.get_attendance_summary()
        if not summary:
            messagebox.showwarning("Warning", "No data to export")
            return
            
        wb = Workbook()
        ws = wb.active
        ws.title = "Attendance Report"
        
        # Add headers
        ws['A1'] = "Attendance Report"
        ws['A2'] = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A4'] = "Name"
        ws['B4'] = "Present Days"
        ws['C4'] = "Total Days"
        ws['D4'] = "Percentage"
        
        # Add data
        for row, item in enumerate(summary, start=5):
            ws.cell(row=row, column=1, value=item['name'])
            ws.cell(row=row, column=2, value=item['present'])
            ws.cell(row=row, column=3, value=item['total'])
            ws.cell(row=row, column=4, value=item['percentage'])
        
        # Save file
        default_path = self.controller.settings["default_save_path"]
        file_path = filedialog.asksaveasfilename(
            initialdir=default_path,
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            title="Save Report"
        )
        
        if file_path:
            try:
                wb.save(file_path)
                messagebox.showinfo("Success", f"Report saved to:\n{file_path}")
            except Exception as e:
                messagebox.showerror("Error", f"Failed to save report:\n{str(e)}")
    
    def generate_report(self):
        """Generate a printable report"""
        self.update_report()
        messagebox.showinfo("Info", "Report generated successfully!")
    
    def print_report(self):
        """Print the report"""
        self.update_report()
        messagebox.showinfo("Info", "Printing would be implemented here")
    
    def on_show(self):
        """Called when the frame is shown"""
        self.update_report()

class SettingsPage(tk.Frame):
    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)
        self.controller = controller
        self.configure(bg="#f0f0f0")
        
        # Header
        header = ttk.Frame(self)
        header.pack(fill="x", pady=10)
        
        ttk.Button(header, text="← Main Menu", 
                  command=lambda: controller.show_frame("MainMenu")).pack(side="left", padx=10)
        
        ttk.Label(header, text="Settings", style="Title.TLabel").pack(side="left", expand=True)
        
        # Content
        self.content = ttk.Frame(self)
        self.content.pack(fill="both", expand=True, padx=20, pady=20)
        
        # Initialize settings display
        self.init_settings_display()
    
    def init_settings_display(self):
        """Initialize the settings display"""
        # Clear existing widgets
        for widget in self.content.winfo_children():
            widget.destroy()
        
        # Settings form
        form_frame = ttk.LabelFrame(self.content, text="Application Settings", padding=10)
        form_frame.pack(fill="x", pady=10)
        
        # Default save path
        ttk.Label(form_frame, text="Default Save Path:").grid(row=0, column=0, sticky="e", padx=5, pady=5)
        self.save_path_var = tk.StringVar(value=self.controller.settings["default_save_path"])
        ttk.Entry(form_frame, textvariable=self.save_path_var, width=40).grid(row=0, column=1, sticky="w", padx=5, pady=5)
        ttk.Button(form_frame, text="Browse...", command=self.browse_save_path).grid(row=0, column=2, padx=5, pady=5)
        
        # Theme selection
        ttk.Label(form_frame, text="Theme:").grid(row=1, column=0, sticky="e", padx=5, pady=5)
        self.theme_var = tk.StringVar(value=self.controller.settings["theme"])
        ttk.Combobox(form_frame, textvariable=self.theme_var, 
                    values=["light", "dark"], state="readonly").grid(row=1, column=1, sticky="w", padx=5, pady=5)
        
        # Save button
        button_frame = ttk.Frame(form_frame)
        button_frame.grid(row=2, column=0, columnspan=3, pady=10)
        ttk.Button(button_frame, text="Save Settings", command=self.save_settings).pack(pady=10)
    
    def browse_save_path(self):
        """Browse for a default save path"""
        path = filedialog.askdirectory(initialdir=self.save_path_var.get())
        if path:
            self.save_path_var.set(path)
    
    def save_settings(self):
        """Save the current settings"""
        self.controller.settings["default_save_path"] = self.save_path_var.get()
        self.controller.settings["theme"] = self.theme_var.get()
        messagebox.showinfo("Success", "Settings saved successfully!")
    
    def on_show(self):
        """Called when the frame is shown"""
        self.save_path_var.set(self.controller.settings["default_save_path"])
        self.theme_var.set(self.controller.settings["theme"])

if __name__ == "__main__":
    app = AttendanceApp()
    app.mainloop()
